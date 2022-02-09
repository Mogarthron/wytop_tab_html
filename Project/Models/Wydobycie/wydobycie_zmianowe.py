import pandas as pd
import json


from Models.Wydobycie.raport_topiarza import Tabelka_z_Wydobyciem_po_baniakach
from Models.Wydobycie.wydobycie_brutto import Wydobycie_Brutto
from Models import scierzki_do_plikow


class Raport_Wydobycia_Zmianowego:
        
    def Raport_Wydobycia_2017(self):
        return pd.read_excel(scierzki_do_plikow["Wytop_2017"])

    def Raport_Wydobycia_2018_2021_11(self):
        rap_2018_2021 = pd.read_excel(scierzki_do_plikow["Dane_Do_Wydobycia"])
        rap_2018_2021.drop(["Unnamed: 8", "Unnamed: 9", "Unnamed: 10"], axis=1, inplace=True)
        return rap_2018_2021




def Wydobycie_Zmianiowe():
    
    df = pd.read_excel(scierzki_do_plikow["Dane_Do_Wydobycia"])

    df.drop(["Unnamed: 8", "Unnamed: 9", "Unnamed: 10"], axis=1, inplace=True)
    
    return df.tail(30)


def Dane_Do_wytopu(od:str, do:str):
    """
    Raport Wytopu na podstawie Raportu Topiarza i Raportera produkcji od 1 listopada 2021
    od,do : format daty RRRR.M.D
    """
    w_brutto = Wydobycie_Brutto(od, do)

    sbdw = w_brutto.Szklo_bezbarwne_do_wytopu 
    wytop_brutto_podsumowanie = w_brutto.Wydobycie_Brutto_Podsumowanie

    def Szklo_bezb_z_produkcji(Zmiana, Mnoznik_Korygujacy_Produkcje=1.2):
        _Zmiana = f"WG_{Zmiana}"
        szklo_bezb_z_produkcji = pd.DataFrame({"Data": pd.date_range(od,do)})
        szklo_bezb_z_produkcji = pd.merge(szklo_bezb_z_produkcji, sbdw.loc[sbdw["Zmiana"] == Zmiana][["Dzm", "Szklo WG"]], how="left", left_on="Data", right_on="Dzm")
        szklo_bezb_z_produkcji.drop("Dzm", axis=1, inplace=True)
        szklo_bezb_z_produkcji["Szklo WG"] = szklo_bezb_z_produkcji["Szklo WG"] * Mnoznik_Korygujacy_Produkcje
        szklo_bezb_z_produkcji.columns = ["Data",_Zmiana] #f"WG_{Zmiana}"]
        szklo_bezb_z_produkcji.fillna(0, inplace=True)
        return szklo_bezb_z_produkcji.groupby(["Data"]).sum()

    
    szklo_bezb_do_wytopu = pd.merge(pd.merge(Szklo_bezb_z_produkcji("R"), Szklo_bezb_z_produkcji("P"), how="inner", on="Data"),
                            Szklo_bezb_z_produkcji("N"),
                            how="inner", on="Data")


    tabeleczka = pd.merge(szklo_bezb_do_wytopu, Tabelka_z_Wydobyciem_po_baniakach(od, do, "2022"), how="inner", on="Data")

    tabeleczka = pd.merge(tabeleczka, Wylewanie(od,do), how="inner", on="Data")

    def Wydobycie_WE(n):
        if type(n["Wylewanie"]) == list and len(n["Wylewanie"]) == 6:
            w_we = 0 if n["Wylewanie"][3] == True else n["WE_1"]
            w_we += 0 if n["Wylewanie"][4] == True else n["WE_2"]
            w_we += 0 if n["Wylewanie"][5] == True else n["WE_3"]

            return w_we

        if type(n["Wylewanie"]) == list and len(n["Wylewanie"]) == 2:
            w_we = 0 if n["Wylewanie"][1] == True else n["WE_1"]
            w_we += 0 if n["Wylewanie"][1] == True else n["WE_2"]
            w_we += 0 if n["Wylewanie"][1] == True else n["WE_3"]

            return w_we

        else:
            return n["WE_1"] + n["WE_2"] + n["WE_3"]


    def Wydobycie_WG(n, Zmiana):

        zmiany = {"R": ["WG_R", "WE_1", 0],
                  "P": ["WG_P", "WE_2", 1],
                  "N": ["WG_N", "WE_3", 2]}

        war = zmiany[Zmiana]

        if type(n["Wylewanie"]) == list and len(n["Wylewanie"]) == 6:
                
            if n["Wylewanie"][war[2]] == True:
                return 0
            elif n["Wylewanie"][war[2]+3] == True:
                return n[war[0]]
            else:
                return n[war[0]] + n[war[1]] * 3 
        
        elif type(n["Wylewanie"]) == list and len(n["Wylewanie"]) == 2:
                
            if n["Wylewanie"][0] == False:
                return 0
            # elif n["Wylewanie"][war[2]+1] == True:
            #     return n[war[0]]
            else:
                return n[war[0]] + n[war[1]] * 3

        elif n["Wylewanie"]:
            return 0
        else:
            return n[war[0]] + n[war[1]] * 3 


    tabeleczka["Wydobycie_WE"] = tabeleczka.apply(lambda x: Wydobycie_WE(x), axis=1)
    tabeleczka["WG_1"] = tabeleczka.apply(lambda x: Wydobycie_WG(x, "R"), axis=1)
    tabeleczka["WG_2"] = tabeleczka.apply(lambda x: Wydobycie_WG(x, "P"), axis=1)
    tabeleczka["WG_3"] = tabeleczka.apply(lambda x: Wydobycie_WG(x, "N"), axis=1)
    tabeleczka["Wydobycie_WG"] = tabeleczka["WG_1"] + tabeleczka["WG_2"] + tabeleczka["WG_3"]
    tabeleczka["Wydobycie_Calkowite"] = tabeleczka["Wydobycie_WG"] + tabeleczka["Wydobycie_WE"]

    tabeleczka = pd.merge(tabeleczka, wytop_brutto_podsumowanie, how="left", left_on="Data", right_on="Dzm")
    tabeleczka.drop("Dzm", axis=1, inplace=True)
    
    # kolumny = ["Data", "WG_1", "WG_2", "WG_3", "Wydobycie_WG", "WE_1", "WE_2", "WE_3", "Wydobycie_WE", "Wydobycie_Calkowite", "Wydobycie_Brutto_Calkowite"]
    # return tabeleczka[kolumny]
    return tabeleczka


def Wylewanie(od,do):

    with open(scierzki_do_plikow["Wylewanie"]) as w:
        dane_Wylewanie = json.load(w)["Zmiany"]        

    wylewanie = pd.DataFrame(dane_Wylewanie)
    wylewanie["Data"] = pd.to_datetime(wylewanie["Data"])
    df = pd.DataFrame({"Data": pd.date_range(od, do)})

    df = pd.merge(df, wylewanie, how="left", on="Data")
    df["Wylewanie"].fillna(False, inplace=True)

    return df


class Drukuj_Raport_Do_Excella:
    # kolumny = ["Data", "WG_1", "WG_2", "WG_3", "Wydobycie_WG", "WE_1", "WE_2", "WE_3", "Wydobycie_WE", "Wydobycie_Calkowite", "Wydobycie_Brutto_Calkowite"]
    
    def __init__(self, od:str, do:str) -> None:
        self.__tabeleczka = Dane_Do_wytopu(od, do)


    def Drukuj_Surowa_Tabelke(self):
        self.__tabeleczka.to_excel("Data\Wytop_listopad_2021.xlsx", index=False)

    def Drukuj_Sformatowny_Raport_Wytopu(self, nazwa_pliku:str):
        """
        nazwa_pliku: nazwa pliku ale też ścieżka dostępu
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Color, Alignment, PatternFill, Border, Side
        
        czerwony_kolor_czcionki_przy_wylewaniu = Font(color="00FF0000")
        _wypelnienie_wiersza = Color(rgb="0099CCFF")

        wb = Workbook()
        ws = wb.active

        # Wypełnienie wierszy tabelki kolorem
        wypelnienie_wiersza = PatternFill(fgColor=_wypelnienie_wiersza, patternType="solid")

        ws["A1"].fill = wypelnienie_wiersza
        ws['B1'].fill = wypelnienie_wiersza
        ws['C1'].fill = wypelnienie_wiersza
        ws['D1'].fill = wypelnienie_wiersza
        ws['E1'].fill = wypelnienie_wiersza
        ws['F1'].fill = wypelnienie_wiersza


        for i in range(self.__tabeleczka.shape[0]):
            if (i%2) == 0:
                ws[f"A{i+3}"].fill = wypelnienie_wiersza
                ws[f"B{i+3}"].fill = wypelnienie_wiersza
                ws[f"C{i+3}"].fill = wypelnienie_wiersza
                ws[f"D{i+3}"].fill = wypelnienie_wiersza
                ws[f"E{i+3}"].fill = wypelnienie_wiersza
                ws[f"F{i+3}"].fill = wypelnienie_wiersza
        
        # Obramowania komorek tabelki 
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

        for i in range(self.__tabeleczka.shape[0]+2):
            for j in range(6):
                ws.cell(row=i+1, column=j+1).border = thin_border
                
        # wiersz podsumowania        
        ws["B1"] = f"Suma WG: {self.__tabeleczka['Wydobycie_WG'].sum():.0f}kg"
        ws["B1"].alignment = Alignment(wrap_text=True, vertical="center")
        ws["C1"] = "WE"        
        ws["F1"] = f"Suma WE: {self.__tabeleczka['Wydobycie_WE'].sum():.0f}kg"
        ws["F1"].alignment = Alignment(wrap_text=True, vertical="center")
        # wiersz zmian
        ws["A2"] = "zmiana"        
        ws["C2"] = "R"
        ws["D2"] = "P"
        ws["E2"] = "N"

        for i in range(self.__tabeleczka.shape[0]):
           
            ws[f"A{i+3}"] = self.__tabeleczka["Data"][i].strftime('%Y-%m-%d')
            ws[f"B{i+3}"] = "" if self.__tabeleczka["Wydobycie_WG"][i] == 0 else self.__tabeleczka["Wydobycie_WG"][i]
            ws[f"C{i+3}"] = self.__tabeleczka["WE_1"][i]     
            ws[f"D{i+3}"] = self.__tabeleczka["WE_2"][i]        
            ws[f"E{i+3}"] = self.__tabeleczka["WE_3"][i]       
            ws[f"F{i+3}"] = "" if self.__tabeleczka["Wydobycie_WE"][i] == 0 else self.__tabeleczka["Wydobycie_WE"][i]

            

        # wylewanie
        for i in range(self.__tabeleczka.shape[0]):
            if type(self.__tabeleczka["Wylewanie"][i]) == list and len(self.__tabeleczka["Wylewanie"][i]) == 2: 
                if self.__tabeleczka["Wylewanie"][i][1] == True:
                    ws[f"C{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                    ws[f"D{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                    ws[f"E{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu 

            if type(self.__tabeleczka["Wylewanie"][i]) == list and len(self.__tabeleczka["Wylewanie"][i]) == 6: 
                
                if self.__tabeleczka["Wylewanie"][i][3] == True and self.__tabeleczka["Wylewanie"][i][4] == True:
                    ws[f"C{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                    ws[f"D{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                elif self.__tabeleczka["Wylewanie"][i][3] == True and self.__tabeleczka["Wylewanie"][i][5] == True:
                    ws[f"C{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                    ws[f"E{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                elif self.__tabeleczka["Wylewanie"][i][4] == True and self.__tabeleczka["Wylewanie"][i][5] == True:
                    ws[f"E{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                    ws[f"D{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                
                if self.__tabeleczka["Wylewanie"][i][3] == True:
                    ws[f"C{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                if self.__tabeleczka["Wylewanie"][i][4] == True:
                    ws[f"D{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu
                if self.__tabeleczka["Wylewanie"][i][5] == True:
                    ws[f"E{i+3}"].font = czerwony_kolor_czcionki_przy_wylewaniu

        wb.save(f"{nazwa_pliku}.xlsx")